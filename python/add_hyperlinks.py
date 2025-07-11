import os
import json
import sys
import urllib.parse
from pathlib import Path
from dotenv import load_dotenv
from openpyxl import load_workbook
from openpyxl.styles import Font

# Load .env from project root
dotenv_path = Path(__file__).resolve().parent.parent / ".env"
load_dotenv(dotenv_path)

excel_path = Path(os.getenv("EXCEL_PATH"))
sharepoint_base_url = os.getenv("SHAREPOINT_BASE_URL")
library_url = os.getenv("SHAREPOINT_LIBRARY_URL")
pdf_folder_base = os.getenv("SHAREPOINT_PDF_FOLDER")

cfg_file = Path(__file__).resolve().parent.parent / "config.json"
if not cfg_file.exists():
    raise FileNotFoundError("config.json not found")
config = json.loads(cfg_file.read_text())
sheet_configs = config["SHEET_CONFIGS"]

# Determine which sheet to process (optional CLI arg)
args = [a for a in sys.argv[1:] if a != "--debug"]
selected_sheet = args[0] if args else None

wb = load_workbook(excel_path)

sheets = [selected_sheet] if selected_sheet else list(sheet_configs.keys())

for sheet_name in sheets:
    cfg = sheet_configs.get(sheet_name)
    if not cfg:
        print(f"⚠️ Sheet '{sheet_name}' not found in config.json", file=sys.stderr)
        continue
    if sheet_name not in wb.sheetnames:
        print(f"⚠️ Sheet not found in workbook: {sheet_name}", file=sys.stderr)
        continue

    ws_original = wb[sheet_name]
    new_sheet = sheet_name.replace(" - Rows", " - Hyperlink")

    if new_sheet in wb.sheetnames:
        wb.remove(wb[new_sheet])
    ws_new = wb.create_sheet(title=new_sheet)

    header = next(ws_original.iter_rows(min_row=1, max_row=1, values_only=True))
    try:
        pdf_col = header.index("PDF Name") + 1
    except ValueError:
        print(f"⚠️ 'PDF Name' column missing in {sheet_name}", file=sys.stderr)
        continue

    ws_new.cell(row=1, column=1, value="PDF Name")
    hyperlink_font = Font(color="0000FF", underline="single")

    encoded_listurl = urllib.parse.quote(library_url, safe='')
    parent_folder = f"{pdf_folder_base}/{cfg['sp_subfolder']}"
    encoded_parent = urllib.parse.quote(parent_folder, safe='')

    for row_idx, row in enumerate(ws_original.iter_rows(min_row=2, values_only=False), start=2):
        cell = row[pdf_col - 1]
        pdf_name = cell.value
        new_cell = ws_new.cell(row=row_idx, column=1, value=pdf_name)
        if not (isinstance(pdf_name, str) and pdf_name.strip()):
            continue
        pdf_name = pdf_name.strip()
        local_pdf = Path(cfg['local_pdf_dir']).resolve() / f"{pdf_name}.pdf"
        if not local_pdf.exists():
            continue
        file_path = f"{pdf_folder_base}/{cfg['sp_subfolder']}/{pdf_name}.pdf"
        encoded_id = urllib.parse.quote(file_path, safe='')
        url = f"{sharepoint_base_url.rstrip('/')}/shared?listurl={encoded_listurl}&id={encoded_id}&parent={encoded_parent}"
        new_cell.hyperlink = url
        new_cell.font = hyperlink_font

wb.save(excel_path)
print(f"✅ Hyperlink sheets updated in: {excel_path}")
